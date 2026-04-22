import time
import re
import pandas as pd
from datetime import datetime
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import requests
import json

# --- CONFIGURATION ---
TICKERS = ["SPY", "QQQ", "MU","NVDA", "SNDK", "AAOI", "ALAB", "TSLA"]
BASE_URL = "https://mztrading.netlify.app/options/analyze/{}?dgextab=GEX&dte=30&showHeatmap=true"
INTERVAL = 900 
FILENAME = "Market_GEX_Heatmaps.xlsx"
SHEETS_BRIDGE_URL = "https://script.google.com/macros/s/AKfycbzu-zhE2_ZKd1hi_YRj4c023BJMpjYOZO_5u54pFPWTZjg_ByssvIxnJ95cmebg1dQl/exec"
# Dictionary to track timestamps in memory for the console output
status_tracker = {ticker: "Not yet fetched" for ticker in TICKERS}

def rgb_to_hex(rgb_str):
    try:
        nums = re.findall(r'\d+', rgb_str)
        if len(nums) >= 3:
            return '{:02X}{:02X}{:02X}'.format(int(nums[0]), int(nums[1]), int(nums[2]))
        return "FFFFFF"
    except:
        return "FFFFFF"

def update_summary_sheet(wb):
    """Creates a 'Summary' tab with all tickers and their last fetch times."""
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        wb.remove(ws)
    
    ws = wb.create_sheet("Summary", 0) # Make it the first tab
    ws.append(["Ticker", "Last Successful Fetch"])
    
    # Style the header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for ticker, timestamp in status_tracker.items():
        ws.append([ticker, timestamp])
    
    # Auto-adjust column width
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25

def scrape_ticker(page, ticker):
    url = f"https://mztrading.netlify.app/options/analyze/{ticker}?dgextab=GEX&dte=30&showHeatmap=true"
    try:
        page.goto(url, wait_until="networkidle", timeout=60000)
        page.wait_for_selector("td", timeout=30000)
        
        rows = page.query_selector_all("tr")
        values_table = []
        colors_table = []

        for row in rows:
            cells = row.query_selector_all("td, th")
            v_row = []
            c_row = []
            for cell in cells:
                v_row.append(cell.inner_text().strip())
                bg = cell.evaluate("el => window.getComputedStyle(el).backgroundColor")
                c_row.append(rgb_to_hex(bg)) # Use your existing rgb_to_hex function
            values_table.append(v_row)
            colors_table.append(c_row)

        # Send data to Google Sheets
        payload = {
            "ticker": ticker,
            "values": values_table,
            "colors": colors_table
        }
        response = requests.post(SHEETS_BRIDGE_URL, json=payload)
        print(f"Sent {ticker} to Sheets: {response.text}")

    except Exception as e:
        print(f"Error {ticker}: {e}")

# ... (Keep all your imports and helper functions like rgb_to_hex)

def run_once():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        context = browser.new_context(user_agent=user_agent)
        page = context.new_page()

        try:
            wb = load_workbook(FILENAME)
        except:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # Create a list to hold data for the CSV
        all_data_frames = []

        for ticker in TICKERS:
            scrape_ticker_to_sheet(page, ticker, wb)
            
            # --- NEW CSV LOGIC START ---
            try:
                import pandas as pd
                ws = wb[ticker]
                # Convert the excel sheet values into a Pandas DataFrame
                data = list(ws.values)
                if len(data) > 0:
                    temp_df = pd.DataFrame(data)
                    # Add a column so you know which ticker this row belongs to
                    temp_df.insert(0, 'Ticker_ID', ticker)
                    all_data_frames.append(temp_df)
            except Exception as e:
                print(f"Could not prepare CSV data for {ticker}: {e}")
            # --- NEW CSV LOGIC END ---

        # Finalize and Save Excel
        update_summary_sheet(wb)
        wb.save(FILENAME)
        
        # --- SAVE THE MASTER CSV ---
        if all_data_frames:
            import pandas as pd
            final_df = pd.concat(all_data_frames, ignore_index=True)
            final_df.to_csv("latest_data.csv", index=False)
            print("Master CSV saved successfully.")

        context.close()
        browser.close()
